import requests
import time,os
import sqlite3,math
from threading import Thread
from datetime import datetime
from queue import Queue
import json
import openpyxl
from openpyxl.styles import Font, PatternFill
import sys
import re,shutil
import win32com.client
import pandas as pd
import zipfile
import logging
import warnings
warnings.filterwarnings("ignore")



abspath = os.path.dirname(os.path.abspath(__file__))

def run_time(func):
    def wrapper(*args, **kw):
        start = time.time()
        func(*args, **kw)
        end = time.time()
        print('耗时：', end-start, 's')
    return wrapper

class Spider():
    def __init__(self,dicMenu):
        self.qpage = Queue()
        self.qurl = Queue()
        self.data = list()
        self.insertdata=list()
        self.getinfo_thread_num = 10
        self.geturl_thread_num = 10
        self.dicMenu=dicMenu
        self.headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) ""Chrome/92.0.4515.107 Safari/537.36 Edg/92.0.902.62",
                        "Content-Type": "application/json","Referer": "https://ecp.sgcc.com.cn/ecp2.0/portal/",}
        
    def produce_pagelist(self):
        for k in self.dicMenu:
            menuID=k
            page_num=self.dicMenu[k]
            '''
            2018032700290425:资格预审公告
            2018032700291334:招标公告及投标邀请书
            2018032900295987:采购公告
            2018060501171107:推荐中标候选人公示
            2018060501171111:中标（成交）结果公告
            2019071434439387:公共信息
            '''
            if menuID=="2018032900295987" or menuID=="2018032700291334" or menuID=="2018032700290425":
                reqdata = '{"index":%d,"size":20,"firstPageMenuId":"%s","purOrgStatus":"","purOrgCode":"","purType":"",''"orgId":"","key":"","orgName":""}'
            else:
                reqdata='{"index":%d,"size":20,"firstPageMenuId":"%s","orgId":"","key":"","year":"","orgName":""}' 
            print("%s一共需要爬取%d页"%(menuID,page_num)) 
            for i in range(page_num , 0,-1):
                datalist = reqdata%(i,menuID)
                #print("%s第%d页"%(menuID,i))
                #print ("生成{%s}URL存入队列,等待其他线程提取"%(menuID))
                self.qpage.put(datalist)# 生成URL存入队列，等待其他线程提取
        print("完成生成连接")         

    def get_info(self):
        notelisturl = 'https://ecp.sgcc.com.cn/ecp2.0/ecpwcmcore//index/noteList' 
        while not self.qpage.empty(): # 保证url遍历结束后能退出线程
            reqdata = self.qpage.get() # 从队列中获取URL
            time.sleep(1)
            req = requests.post(url=notelisturl, headers=self.headers,data=reqdata)
            data = json.loads(req.text)  
            req.close()
            if data['successful']==True:
                lensize=len(data['resultValue']["noteList"])
                for i in range(lensize-1, -1, -1):
                    reslut ={
                    'noticid': data['resultValue']["noteList"][i]['id'],
                    'doctype': data['resultValue']["noteList"][i]['doctype'],
                    'menuID': data['resultValue']["noteList"][i]["firstPageMenuId"]
                    }
                    self.data.append(reslut)
            else:
                print ("获取URL失败")
        print("成功完成提取URL")

    def WXworkHook(self,sendurl,path,filename,sheet_name):
        wxhookurl="https://qyapi.weixin.qq.com/cgi-bin/webhook/send?key=919183d0-925c-4872-b122-d2e079099a8d"
        #wxhookurl="https://qyapi.weixin.qq.com/cgi-bin/webhook/send?key=f16befea-b333-4384-8793-4d027324542f"
        mHeader = {'Content-Type': 'application/json;charset=utf-8'}
        msg="""
        **新增国网低压开关柜招标**
        >[{url}]({url})
        >Excel文件：{name}
        >Sheet名称：{sheet}
        """.format(url=sendurl,name=filename,sheet=sheet_name)    
        data = {"msgtype": "markdown", "markdown": {"content": msg}}
        r = requests.post(wxhookurl, headers=mHeader, json=data)
        print (msg)
        
    def zippd(self,url,fpath):
        key_set=os.path.join(abspath, 'keywords.xlsx')
        keywords_df = pd.read_excel(key_set,usecols=['物料'])
        params_set = set(keywords_df['物料'].dropna().astype(str))   
        savefile=os.path.join(abspath, '标准化柜.xlsx')
        with zipfile.ZipFile(fpath, 'r') as zip_ref:
            files=zip_ref.namelist()
            for file in files:
                if file.endswith('.xlsx') and file.__contains__("货物清单"):
                    with zip_ref.open(file, 'r') as f:
                        dfexcel = pd.read_excel(f,sheet_name=None)
                        for sheet_name, df in dfexcel.items():
                            dftemp=pd.DataFrame(df)
                            mask = dftemp['物料编码'].astype(str).isin(params_set)        
                            matching_rows = df[mask]
                            if not os.path.exists(savefile):
                                empty_df = pd.DataFrame()
                                empty_df = matching_rows.copy()
                                empty_df.to_excel(savefile, index=False)
                            else:
                                if not matching_rows.empty:
                                    #msg=datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")+"  找到匹配的文件:PATH:{},   filename:{},    sheet_name:{}".format(fpath,file,sheet_name)
                                    self.WXworkHook(url,fpath,file,sheet_name)
                                    logging.info("找到匹配的文件:PATH:%s,   filename:%s,    sheet_name:%s",fpath,file,sheet_name)
                                    with pd.ExcelWriter(savefile, mode='a', engine='openpyxl', if_sheet_exists='overlay') as writer:
                                        matching_rows.to_excel(writer, sheet_name='Sheet1', index=False,header=None, startrow=writer.sheets['Sheet1'].max_row)          
                                    df = pd.read_excel(savefile)
                                    df_unique = df.drop_duplicates()
                                    df_unique.to_excel(savefile, index=False)


    def getDownloadUrl(self,bidurl,__filename,bidorg):
        __link='https://ecp.sgcc.com.cn/ecp2.0/ecpwcmcore//index/downLoadBid?noticeId='+str(__filename)+'&noticeDetId='   
        create_dir_not_exist(abspath + '\\Download\\'+str(bidorg))    
        dst=abspath + '\\Download\\' + str(bidorg)+"\\"+str(__filename) + '.zip'
        if not os.path.exists(dst) or os.path.getsize(dst)==0:
            try:
        # Open Connection
                Filedown = requests.get(__link, headers=self.headers,stream=True)
            except Exception as e:
                print(e)
                print("错误，访问url: %s 异常" % __link)       
            with open(dst,'wb+') as f:
            # 分块写入文件
                for chunk in Filedown.iter_content(chunk_size=1024):
                    if chunk:
                        f.write(chunk)
            f.close()
            print(datetime.now().strftime("%Y-%m-%d %H:%M:%S")+'\t'+__link+"\t文件下载成功")
            self.zippd(bidurl,dst)
        else:
            print (datetime.now().strftime("%Y-%m-%d %H:%M:%S") +'\t'+__link+"\t文件已存在")
        return '=HYPERLINK("'+dst+ '","' +str(__filename) + '.zip' + '")'

  
    def insert_data(self):
        conn = sqlite3.connect(abspath + '\\Res\\sgcc.db')
        try:
            # 1.创建游标对象
            cursor = conn.cursor()
            for d in self.insertdata:
                columns = ', '.join(d["data"].keys())
                placeholders = ':' + ', :'.join(d["data"].keys())
                query = 'INSERT INTO hisdata (%s) select %s  where not exists (select 1 from hisdata where noticeId=%s)' % (columns, placeholders,d["noticid"])
                cursor.execute(query, d["data"])
            conn.commit()
            efrow=cursor.lastrowid
            return efrow
        except sqlite3.DatabaseError as error:
            print(error)
            conn.rollback()
        finally:
            conn.close()   
            
    def parseURL(self,lstid, doctype, menuid):
        Url = 'https://ecp.sgcc.com.cn/ecp2.0/ecpwcmcore//index/'
        bidlist = []
        key = ['gettime', 'noticeId', 'TYPE_NAME']
        typename = ""
        if menuid == 2018032900295987:
            typename = "采购公告"
        elif menuid == 2018032700291334:
            typename = "招标公告"
        elif menuid == 2018032700290425:
            typename = "资格预审公告"
        elif menuid == 2018060501171111:
            typename = "中标（成交）结果公告"
        elif menuid == 2018060501171107:
            typename = "推荐中标候选人公示"
        while True:
            try:
                if doctype == "doci-bid":
                    getNoticeBidurl = Url + 'getNoticeBid'
                    resd = requests.post(url=getNoticeBidurl, data=json.dumps(lstid), headers=self.headers)
                    res = json.loads(resd.text)
                    resd.close()
                    PurType = res["resultValue"]["notice"]["PUR_TYPE_NAME"]
                    if PurType == "物资":
                        bidlist.append(datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
                        bidlist.append(res["resultValue"]["notice"]["PURPRJ_NAME"])
                        bidlist.append(res["resultValue"]["notice"]["PUB_TIME"])
                        bidurl='https://ecp.sgcc.com.cn/ecp2.0/portal/#/doc/doci-bid/' + str(lstid) + '_' + str(menuid)
                        if menuid == 2018032900295987:#采购公告
                            key = ['gettime', 'PURPRJ_NAME', 'PUB_TIME', 'PURPRJ_CODE', 'BIDBOOK_SELL_BEGIN_TIME',
                                'OPENBID_TIME','BIDBOOK_BUY_END_TIME', 'URL', 'BID_AGT','histype', 'ORGNAME', 'noticeId', 'TYPE_NAME']
                            bidlist.append(res["resultValue"]["notice"]["PURPRJ_CODE"])
                            bidlist.append(res["resultValue"]["notice"]["BIDBOOK_SELL_BEGIN_TIME"])
                            bidlist.append(res["resultValue"]["notice"]["OPENBID_TIME"])
                            bidlist.append(res["resultValue"]["notice"]["BIDBOOK_BUY_END_TIME"])
                        elif menuid == 2018032700291334:#招标公告                        
                            key = ['gettime', 'PURPRJ_NAME', 'PUB_TIME', 'PURPRJ_CODE', 'BIDBOOK_SELL_BEGIN_TIME',
                                'OPENBID_TIME','BIDBOOK_BUY_END_TIME', 'zip', 'URL','BID_AGT','histype', 'ORGNAME',  'noticeId', 'TYPE_NAME']
                            bidlist.append(res["resultValue"]["notice"]["PURPRJ_CODE"])
                            bidlist.append(res["resultValue"]["notice"]["BIDBOOK_SELL_BEGIN_TIME"])
                            bidlist.append(res["resultValue"]["notice"]["OPENBID_TIME"])
                            bidlist.append(res["resultValue"]["notice"]["BIDBOOK_BUY_END_TIME"])
                            if res["resultValue"]["fileFlag"]=="1":
                                zipurl = self.getDownloadUrl(bidurl,lstid,res["resultValue"]["notice"]["BID_ORG"])
                                bidlist.append(zipurl)
                            else:
                                bidlist.append('=HYPERLINK("'+bidurl+ '","' +'无附件，见网页' + '")') 
                        elif menuid == 2018032700290425:#资格预审公告
                            key = ['gettime', 'PURPRJ_NAME', 'PUB_TIME', 'URL', 'BID_AGT','histype', 'ORGNAME', 'noticeId', 'TYPE_NAME']
                        bidlist.append(bidurl)
                        bidlist.append(res["resultValue"]["notice"]["BID_AGT"])
                        bidlist.append('0')
                        bidlist.append(res["resultValue"]["notice"]["BID_ORG"])
                        bidlist.append(str(lstid))
                        bidlist.append(typename)
                        insertdata = zip(key, bidlist)
                        parsedata ={
                            'noticid': lstid,
                            'data': dict(insertdata)
                            }
                        self.insertdata.append(parsedata)
                        print(datetime.now().strftime("%Y-%m-%d %H:%M:%S") + "\t新增" + typename.ljust(11) + "\t" + str(lstid))  
                    # else:
                    #     key = ['gettime', 'histype', 'ORGNAME', 'noticeId', 'TYPE_NAME']
                    #     bidlist.append(datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
                    #     bidlist.append('0')
                    #     bidlist.append(res["resultValue"]["notice"]["BID_ORG"])
                elif doctype == "doci-change":
                    getChangeBidurl = Url + 'getChangeBid'
                    resd=requests.post(url=getChangeBidurl, data=json.dumps(lstid), headers=self.headers)
                    res = json.loads(resd.text)
                    resd.close()
                    PurType = res["resultValue"]["origNotice"]["PUR_TYPE_NAME"]
                    bidlist.append(datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
                    if PurType == "物资":
                        bidlist.append(res["resultValue"]["origNotice"]["PURPRJ_NAME"])
                        bidlist.append(res["resultValue"]["origNotice"]["PUB_TIME"])
                        ziporgurl='https://ecp.sgcc.com.cn/ecp2.0/portal/#/doc/doci-change/' + str(lstid) + '_' + str(menuid)
                        if res["resultValue"]["chgNotice"] is not None:
                            if menuid == 2018032900295987:
                                key = ['gettime', 'PURPRJ_NAME', 'PUB_TIME', 'PURPRJ_CODE', 'BIDBOOK_SELL_BEGIN_TIME',
                                    'OPENBID_TIME','BIDBOOK_BUY_END_TIME', 'URL', 'BID_AGT', 'histype', 'ORGNAME', 'noticeId', 'TYPE_NAME']
                                bidlist.append(res["resultValue"]["chgNotice"]["PURPRJ_CODE"])
                                bidlist.append(res["resultValue"]["chgNotice"]["BIDBOOK_SELL_BEGIN_TIME"])
                                bidlist.append(res["resultValue"]["chgNotice"]["OPENBID_TIME"])
                                bidlist.append(res["resultValue"]["chgNotice"]["BIDBOOK_BUY_END_TIME"])
                            elif menuid == 2018032700290425:
                                key = ['gettime', 'PURPRJ_NAME', 'PUB_TIME', 'URL', 'BID_AGT', 'histype', 'ORGNAME', 'noticeId', 'TYPE_NAME']
                            elif menuid == 2018032700291334:
                                key = ['gettime', 'PURPRJ_NAME', 'PUB_TIME', 'PURPRJ_CODE', 'BIDBOOK_SELL_BEGIN_TIME',
                                    'OPENBID_TIME','BIDBOOK_BUY_END_TIME', 'zip', 'URL', 'BID_AGT', 'histype', 'ORGNAME', 'noticeId', 'TYPE_NAME']
                                bidlist.append(res["resultValue"]["chgNotice"]["PURPRJ_CODE"])
                                bidlist.append(res["resultValue"]["chgNotice"]["BIDBOOK_SELL_BEGIN_TIME"])
                                bidlist.append(res["resultValue"]["chgNotice"]["OPENBID_TIME"])
                                bidlist.append(res["resultValue"]["chgNotice"]["BIDBOOK_BUY_END_TIME"])
                                if res["resultValue"]["fileFlag"]=="1":
                                    zipurl =self.getDownloadUrl(ziporgurl,lstid,res["resultValue"]["origNotice"]["BID_ORG"])
                                    bidlist.append(zipurl)
                                else:
                                    bidlist.append('=HYPERLINK("'+ziporgurl+ '","' +'无附件，见网页' + '")') 
                        else:
                            if menuid == 2018032900295987:
                                key = ['gettime', 'PURPRJ_NAME', 'PUB_TIME', 'PURPRJ_CODE', 'BIDBOOK_SELL_BEGIN_TIME',
                                    'OPENBID_TIME','BIDBOOK_BUY_END_TIME', 'URL', 'BID_AGT', 'histype', 'ORGNAME', 'noticeId', 'TYPE_NAME']
                                bidlist.append(res["resultValue"]["origNotice"]["PURPRJ_CODE"])
                                bidlist.append(res["resultValue"]["origNotice"]["BIDBOOK_SELL_BEGIN_TIME"])
                                bidlist.append(res["resultValue"]["origNotice"]["OPENBID_TIME"])
                                bidlist.append(res["resultValue"]["origNotice"]["BIDBOOK_BUY_END_TIME"])
                            elif menuid == 2018032700290425:
                                key = ['gettime', 'PURPRJ_NAME', 'PUB_TIME', 'URL', 'BID_AGT', 'histype', 'ORGNAME', 'noticeId', 'TYPE_NAME']
                            elif menuid == 2018032700291334:
                                key = ['gettime', 'PURPRJ_NAME', 'PUB_TIME', 'PURPRJ_CODE', 'BIDBOOK_SELL_BEGIN_TIME',
                                    'OPENBID_TIME','BIDBOOK_BUY_END_TIME', 'zip', 'URL', 'BID_AGT', 'histype', 'ORGNAME', 'noticeId', 'TYPE_NAME']
                                bidlist.append(res["resultValue"]["origNotice"]["PURPRJ_CODE"])
                                bidlist.append(res["resultValue"]["origNotice"]["BIDBOOK_SELL_BEGIN_TIME"])
                                bidlist.append(res["resultValue"]["origNotice"]["OPENBID_TIME"])
                                bidlist.append(res["resultValue"]["origNotice"]["BIDBOOK_BUY_END_TIME"])
                                if res["resultValue"]["fileFlag"]=="1":
                                    zipurl =self.getDownloadUrl(ziporgurl,lstid,res["resultValue"]["origNotice"]["BID_ORG"])
                                    bidlist.append(zipurl)
                                else:
                                    bidlist.append('=HYPERLINK("'+ziporgurl+ '","' +'无附件，见网页' + '")') 
                        bidlist.append(ziporgurl)
                        bidlist.append(res["resultValue"]["origNotice"]["BID_AGT"])
                        bidlist.append('1')
                        bidlist.append(res["resultValue"]["origNotice"]["BID_ORG"])
                        bidlist.append(str(lstid))
                        bidlist.append(typename)
                        insertdata = zip(key, bidlist)
                        parsedata ={
                            'noticid': lstid,
                            'data': dict(insertdata)
                            }
                        self.insertdata.append(parsedata)
                        print(datetime.now().strftime("%Y-%m-%d %H:%M:%S") + "\t新增" + typename.ljust(11) + "\t" + str(lstid)) 
                elif doctype == "doci-win":
                    getNoticeWinurl = Url + 'getNoticeWin'
                    resd=requests.post(url=getNoticeWinurl, data=json.dumps(lstid), headers=self.headers)
                    key = ['gettime', 'PURPRJ_NAME', 'PUB_TIME', 'URL','BID_AGT','histype', 'ORGNAME', 'noticeId', 'TYPE_NAME','zip']
                    resd.close()
                    res = json.loads(resd.text)
                    bidlist.append(datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
                    bidlist.append(res["resultValue"]["notice"]["TITLE"])
                    if res["resultValue"]["notice"]["PUB_TIME"] is None:
                        bidlist.append(res["resultValue"]["notice"]["UPD_TIME"])
                    else:
                        bidlist.append(res["resultValue"]["notice"]["PUB_TIME"])
                    orgurl='https://ecp.sgcc.com.cn/ecp2.0/portal/#/doc/doci-win/' + str(lstid) + '_' + str(menuid)
                    bidlist.append(orgurl)
                    bidlist.append(res["resultValue"]["notice"]["bidagtName"])
                    bidlist.append('0')
                    bidlist.append(res["resultValue"]["notice"]["bidOrgName"])
                    bidlist.append(str(lstid))
                    bidlist.append(typename)
                    if  res["resultValue"]["fileFlag"]=="1":
                        if res["resultValue"]["notice"]["HIDE_FLAG"] is None:
                            zippath=self.getWinFile(lstid,res["resultValue"]["notice"]["bidOrgName"],orgurl)
                            while True:
                                try:
                                    if zippath:
                                            bidlist.append(zippath)
                                            break
                                except:
                                    print('download error')
                        else:
                            bidlist.append('=HYPERLINK("'+orgurl+ '","' +'公示过期，见网页' + '")')
                    else:
                        bidlist.append('=HYPERLINK("'+orgurl+ '","' +'公示无附件，见网页' + '")')                                        
                    insertdata = zip(key, bidlist)
                    parsedata ={
                        'noticid': lstid,
                        'data': dict(insertdata)
                        }
                    self.insertdata.append(parsedata)
                    print(datetime.now().strftime("%Y-%m-%d %H:%M:%S") + "\t新增" + typename.ljust(11) + "\t" + str(lstid)) 
                elif doctype == "doc-com":
                    getDocUrl = Url + 'getDoc'
                    key = ['gettime', 'PURPRJ_NAME', 'PUB_TIME', 'zip','URL', 'BID_AGT','histype', 'ORGNAME', 'noticeId', 'TYPE_NAME']
                    resd = requests.post(url=getDocUrl, data=json.dumps(lstid), headers=self.headers)
                    res = json.loads(resd.text)
                    resd.close()
                    bidlist.append(datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
                    bidlist.append(res["resultValue"]["doc"]["title"])
                    bidlist.append(res["resultValue"]["doc"]["noticePublishTime"])
                    orgurlcom='https://ecp.sgcc.com.cn/ecp2.0/portal/#/doc/doc-com/' + str(lstid) + '_' + str(menuid)
                    if res["resultValue"]["doc"]["fileFlag"]=="true" and len(res["resultValue"]["files"])>0:
                        filename=res["resultValue"]["files"][0]["fileName"]
                        downpath=res["resultValue"]["files"][0]["filePath"]+'/'+res["resultValue"]["files"][0]["attachName"]
                        zippath=self.getdocFile(downpath,str(lstid)+filename,res["resultValue"]["doc"]["publishOrgName"])
                        while True:
                            try:
                                if zippath:
                                    bidlist.append(zippath)
                                    break
                            except:
                                print('download error')
                    else:
                        bidlist.append('=HYPERLINK("'+orgurlcom+ '","' +'公示无附件，见网页' + '")')              
                    bidlist.append(orgurlcom)
                    bidlist.append(res["resultValue"]["doc"]["noticeInscribeOrg"])
                    bidlist.append('0')
                    bidlist.append(res["resultValue"]["doc"]["publishOrgName"])
                    bidlist.append(str(lstid))
                    bidlist.append(typename)
                    insertdata = zip(key, bidlist)
                    parsedata ={
                        'noticid': lstid,
                        'data': dict(insertdata)
                        }
                    self.insertdata.append(parsedata)
                    print(datetime.now().strftime("%Y-%m-%d %H:%M:%S") + "\t新增" + typename.ljust(11) + "\t" + str(lstid)) 
                elif doctype == "doc-spec":
                    getDocurl = Url + 'getDoc'
                    resd=requests.post(url=getDocurl, data=json.dumps(lstid), headers=self.headers)
                    res = json.loads(resd.text)
                    resd.close()
                    key = ['gettime', 'PURPRJ_NAME', 'PUB_TIME', 'URL', 'histype', 'ORGNAME', 'noticeId', 'TYPE_NAME']
                    bidlist.append(datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
                    bidlist.append(res["resultValue"]["doc"]["purOrgName"])
                    bidlist.append(res["resultValue"]["doc"]["noticePublishTime"])
                    bidlist.append('https://ecp.sgcc.com.cn/ecp2.0/portal/#/doc/doc-spec/' + str(lstid) + '_' + str(menuid))
                    bidlist.append('0')
                    bidlist.append(res["resultValue"]["doc"]["publishOrgName"])
                    bidlist.append(str(lstid))
                    bidlist.append(typename)
                    insertdata = zip(key, bidlist)
                    parsedata ={
                        'noticid': lstid,
                        'data': dict(insertdata)
                        }
                    self.insertdata.append(parsedata)
                    print(datetime.now().strftime("%Y-%m-%d %H:%M:%S") + "\t新增" + typename.ljust(11) + "\t" + str(lstid)) 
            except (requests.exceptions.SSLError, requests.exceptions.ConnectionError) as e:
                if 'bad handshake' in str(e) or '10054'or '10060' in str(e):  # 上述2种异常
                    print("error,bad handshake,10054,10060,wait 2s,continue")
                    time.sleep(2)
                    continue  # 继续发请求
                else:
                    print ('https://ecp.sgcc.com.cn/ecp2.0/portal/#/doc/'+doctype+'/'+str(lstid) + '_' + str(menuid))
                    raise Exception(e)  # 其他异常，抛出来
            break 

        
    def produce_urllist(self):
        for k in self.data:
            self.qurl.put(k) # 生成URL存入队列，等待其他线程提取
            
    def get_parseURL(self):
        while not self.qurl.empty(): # 保证url遍历结束后能退出线程
            t = self.qurl.get() # 从队列中获取URL
            try:
                self.parseURL(t["noticid"], t["doctype"], t["menuID"])
            except:
                print (t["noticid"], t["doctype"], t["menuID"])
            
    def getWinFile(self,noticid,orgname,orgurl):
        getwinfileurl='https://ecp.sgcc.com.cn/ecp2.0/ecpwcmcore//index/getWinFile'
        resd=requests.post(url=getwinfileurl, data=json.dumps(noticid), headers=self.headers)
        res = json.loads(resd.text)
        if res['successful']==True :
            if len(res['resultValue']['files'])<2:
                Fpath=res['resultValue']['files'][0]['FILE_PATH']
                pdf__link='https://ecp.sgcc.com.cn/ecp2.0/ecpwcmcore//index/showPDF?filePath='+Fpath+'&a=b'
                filename=res['resultValue']['files'][0]['FILE_NAME']
                create_dir_not_exist(abspath + '\\PDF\\' + str(orgname))
                dst=abspath + '\\PDF\\' + str(orgname)+"\\" + str(noticid)+str(filename)
                if not os.path.exists(dst) or os.path.getsize(dst)==0:
                    try:
                # Open Connection
                        Filedown = requests.get(pdf__link, headers=self.headers,stream=True)
                    except Exception as e:
                        print(e)
                        print("错误，访问url: %s 异常" % pdf__link)
                        return False
                    with open(dst,'wb+') as f:
                # 分块写入文件
                        for chunk in Filedown.iter_content(chunk_size=1024):
                            if chunk:
                                f.write(chunk)
                    f.close()
                    print(datetime.now().strftime("%Y-%m-%d %H:%M:%S") +'\t'+pdf__link+"\t文件下载成功")
                else:
                    print (datetime.now().strftime("%Y-%m-%d %H:%M:%S")+'\t' +pdf__link+"\t文件已存在")
                return '=HYPERLINK("'+dst+ '","' +str(filename) + '")'
            else:
                return '=HYPERLINK("'+orgurl+ '","' +'有2个附件，见网页' + '")'

    def getdocFile(self,Fpath,__filename,orgname):
        pdf__link='https://ecp.sgcc.com.cn/ecp2.0/ecpwcmcore/index/showPDF?filePath='+Fpath+'&a=b'
        create_dir_not_exist(abspath + '\\PDF\\' + str(orgname))
        dst=abspath + '\\PDF\\' + str(orgname)+"\\" + str(__filename)
        if not os.path.exists(dst) or os.path.getsize(dst)==0:
            try:
            # Open Connection
                Filedown = requests.get(pdf__link, headers=self.headers,stream=True)
            except Exception as e:
                print(e)
                print("错误，访问url: %s 异常" % pdf__link)
                return False
            with open(dst,'wb+') as f:
                    # 分块写入文件
                for chunk in Filedown.iter_content(chunk_size=1024):
                    if chunk:
                        f.write(chunk)
            f.close()
            print(datetime.now().strftime("%Y-%m-%d %H:%M:%S") +'\t'+pdf__link+"\t'文件下载成功")
        else:
            print(datetime.now().strftime("%Y-%m-%d %H:%M:%S")+'\t' +pdf__link+"\t'文件已存在")
        return '=HYPERLINK("'+dst+ '","' +str(__filename) + '")'               
                    
    @run_time
    def run(self):
        self.produce_pagelist()
        ths = []
        thread_list=[]  
        for _ in range(self.getinfo_thread_num):
            th = Thread(target=self.get_info)           
            th.start()
            ths.append(th)  
        for th in ths:
            th.join()
        self.produce_urllist()
        for _ in range(self.geturl_thread_num):
            m = Thread(target=self.get_parseURL)
            m.start()
            thread_list.append(m)
        for m in thread_list:
            m.join()
        print('获取数据结束.开始写入数据库...')
        self.insert_data()
        print('数据库写入完成....')     

def geturlnum():
    print(datetime.now().strftime("%Y-%m-%d %H:%M:%S") + "\t开始获取数量")
    headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) ""Chrome/92.0.4515.107 Safari/537.36 Edg/92.0.902.62","Content-Type": "application/json","Referer": "https://ecp.sgcc.com.cn/ecp2.0/portal/",}
    url = 'https://ecp.sgcc.com.cn/ecp2.0/ecpwcmcore//index/noteList'
    IDlist=[2018032900295987,2018032700291334,2018032700290425,2018060501171107,2018060501171111]
    menudict={}
    sgcc_session  =  requests.Session()
    for menuID in IDlist:
        if menuID==2018032900295987 or menuID==2018032700291334 or menuID==2018032700290425:
            data = '{"index":1,"size":20,"firstPageMenuId":"%s","purOrgStatus":"","purOrgCode":"","purType":"",''"orgId":"","key":"","orgName":""}'
        else:
            data='{"index":1,"size":20,"firstPageMenuId":"%s","orgId":"","key":"","year":"","orgName":""}'  
        reqdata=data%(menuID)
        while True:
            try:
                res = sgcc_session.post(url=url, data=reqdata,headers=headers)
                res.close()
                break
            except sgcc_session.exceptions.ConnectionError:
                print('ConnectionError -- please wait 3 seconds')
                time.sleep(3)
            except sgcc_session.exceptions.ChunkedEncodingError:
                print('ChunkedEncodingError -- please wait 3 seconds')
                time.sleep(3)
            except:
                print('Unfortunitely -- An Unknow Error Happened, Please wait 3 seconds')
                time.sleep(3)
        result = json.loads(res.text)

        if result['successful']!=False:
            num=result['resultValue']['count']
            menudict[str(menuID)]=str(num)
    print(datetime.now().strftime("%Y-%m-%d %H:%M:%S") + "\t获取数量完毕")
    return menudict

def writexls(num):
    try:
        excel =  win32com.client.gencache.EnsureDispatch('Excel.Application')
    except AttributeError:
        MODULE_LIST = [m.__name__ for m in sys.modules.values()]
        for module in MODULE_LIST:
            if re.match(r'win32com\.gen_py\..+', module):
                del sys.modules[module]  
        shutil.rmtree(os.path.join(os.environ.get('LOCALAPPDATA'), 'Temp', 'gen_py'))
        excel = win32com.client.gencache.EnsureDispatch('Excel.Application')  
    time.sleep(3)
    for wb in excel.Workbooks: 
        if wb.Name =='国网招采信息2024.xlsx':
            wb.Close(True)
            print("关闭 excel!")
    if  os.path.exists(abspath + '\\国网招采信息2024.xlsx'):
        pass
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '采购公告'
        wb.save(abspath + '\\国网招采信息2024.xlsx')
        wb.close()
    ws_open = openpyxl.load_workbook(abspath + '\\国网招采信息2024.xlsx')
    lstw = DB2Excel(num)
    for one_bid in lstw:
        bid = [i for i in one_bid[0:11] if i is not None]
        sheet_name = one_bid[-1]
        pattern = one_bid[-2]
        if sheet_name not in ws_open.sheetnames:
            ws_open.create_sheet(sheet_name)
        ws = ws_open[sheet_name]
        if sheet_name == "采购公告" and ws['A1'].value is None:
            ws['A1'] = '获取时间'
            ws.column_dimensions['A'].width = 21
            ws['B1'] = '采购项目名称'
            ws.column_dimensions['B'].width = 97
            ws['C1'] = '采购项目编号'
            ws.column_dimensions['C'].width = 21
            ws['D1'] = '公告发布时间'
            ws.column_dimensions['D'].width = 13
            ws['E1'] = '招标文件获取截止时间'
            ws.column_dimensions['E'].width = 22
            ws['F1'] = '开标（截标）时间'
            ws.column_dimensions['F'].width = 22
            ws['G1'] = '开启应答文件时间'
            ws.column_dimensions['G'].width = 23
            ws['H1'] = '网页链接'
            ws.column_dimensions['H'].width = 90
            ws['I1'] = '采购单位'
            ws.column_dimensions['I'].width = 30
            ws['J1'] = '招标代理机构'
            ws.column_dimensions['J'].width = 30
        elif sheet_name == "招标公告" and ws['A1'].value is None:
            ws['A1'] = '获取时间'
            ws.column_dimensions['A'].width = 21
            ws['B1'] = '招标项目名称'
            ws.column_dimensions['B'].width = 97
            ws['C1'] = '招标项目编号'
            ws.column_dimensions['C'].width = 21
            ws['D1'] = '公告发布时间'
            ws.column_dimensions['D'].width = 13
            ws['E1'] = '招标文件获取截止时间'
            ws.column_dimensions['E'].width = 22
            ws['F1'] = '开标（截标）时间'
            ws.column_dimensions['F'].width = 22
            ws['G1'] = '开启应答文件时间'
            ws.column_dimensions['G'].width = 23
            ws['H1'] = '公告文件'
            ws.column_dimensions['H'].width = 22
            ws['I1'] = '网页链接'
            ws.column_dimensions['I'].width = 90
            ws['J1'] = '采购单位'
            ws.column_dimensions['J'].width = 30
            ws['K1'] = '招标代理机构'
            ws.column_dimensions['K'].width = 30
        elif sheet_name == "中标（成交）结果公告" or sheet_name =="推荐中标候选人公示" and ws['A1'].value is None:
            ws['A1'] = '获取时间'
            ws.column_dimensions['A'].width = 21
            ws['B1'] = '项目名'
            ws.column_dimensions['B'].width = 97
            ws['C1'] = '发布时间'
            ws.column_dimensions['C'].width = 21
            ws['D1'] = '公告文件'
            ws.column_dimensions['D'].width = 80
            ws['E1'] = '网页链接'
            ws.column_dimensions['E'].width = 90
            ws['F1'] = '采购单位'
            ws.column_dimensions['F'].width = 30
            ws['G1'] = '招标代理机构'
            ws.column_dimensions['F'].width = 30            
        elif sheet_name == "资格预审公告" and ws['A1'].value is None:
            ws['A1'] = '获取时间'
            ws.column_dimensions['A'].width = 21
            ws['B1'] = '项目名'
            ws.column_dimensions['B'].width = 97
            ws['C1'] = '发布时间'
            ws.column_dimensions['C'].width = 21
            ws['D1'] = '网页链接'
            ws.column_dimensions['D'].width = 90
            ws['E1'] = '采购单位'
            ws.column_dimensions['E'].width = 30
            ws['F1'] = '招标代理机构'
            ws.column_dimensions['F'].width = 30                
        if len(bid) > 5:
            if(sheet_name == "招标公告"):
                ws.append(one_bid[0:11])
            else:
                ws.append(bid)
            if pattern == '1':
                color_fill = PatternFill(fill_type='solid', fgColor="AACF91")
                for j in range(1, ws.max_column + 1):
                    ws.cell(row=ws.max_row, column=j).fill = color_fill
            elif "库存"  in bid[1]:
                font = Font(bold=True, color="FF0000")
                ws.cell(row=ws.max_row, column=2).font = font
            if "物资" in bid[1] and "非物资" not in bid[1]:
                color_fill = PatternFill(fill_type='solid', fgColor="FFD700")
                for j in range(1, ws.max_column + 1):
                    ws.cell(row=ws.max_row, column=j).fill = color_fill
    print(datetime.now().strftime("%Y-%m-%d %H:%M:%S") + "\twrite to excel!")
    ws_open.save(abspath + '\\国网招采信息2024.xlsx')
    ws_open.close()
    # excel.Workbooks.Open(abspath + '\\国网招采信息2024.xlsx', ReadOnly = False) 
    # excel.Visible = True  

#增量读取数据库
def DB2Excel(num):
#def DB2Excel():
    conn = sqlite3.connect(abspath + '\\Res\\sgcc.db')
    towrite = []
    try:
        # 1.创建游标对象
        cursor = conn.cursor()
        # 2.执行SQL操作
        cursor.execute("select * from (select * from hisdata order by id desc limit 0,?)as tbl order by id ", (num,))
        #cursor.execute("select * from (select * from hisdata order by id desc )as tbl order by id ")
        data = cursor.fetchall()
        if data is not None:
            for i in data:
                read_db = list(i[1:15])
                towrite.append(read_db)
            return towrite
        else:
            return 0
    except sqlite3.DatabaseError as error:
        # 4. 回滚数据库事物
        print(error)
        conn.rollback()
    finally:
        # 5. 关闭数据连接
        conn.close()

       
def create_table():
    conn = sqlite3.connect(abspath + '\\Res\\sgcc.db')
    try:
        create_tb_cmd = '''
        CREATE TABLE IF NOT EXISTS hisdata (
        id                      INTEGER PRIMARY KEY AUTOINCREMENT
                                        NOT NULL,
        gettime                 TEXT,
        PURPRJ_NAME             TEXT,
        PURPRJ_CODE             TEXT,
        PUB_TIME                TEXT,
        BIDBOOK_SELL_BEGIN_TIME TEXT,
        OPENBID_TIME            TEXT,
        BIDBOOK_BUY_END_TIME    TEXT,
        zip                     TEXT,
        URL                     TEXT,
        ORGNAME                 TEXT,
        BID_AGT                 TEXT,
        noticeId                TEXT,
        histype                 TEXT,
        TYPE_NAME        TEXT);
        '''
        createinfo_cmd='''
        CREATE TABLE IF NOT EXISTS info (
        id                 INTEGER PRIMARY KEY AUTOINCREMENT
                                        NOT NULL,
        [2018032900295987] TEXT,
        [2018032700291334] TEXT,
        [2018032700290425] TEXT,
        [2018060501171111] TEXT,
        [2018060501171107] TEXT,
        runtime            TEXT);
        '''
        conn.execute(create_tb_cmd)
        conn.execute(createinfo_cmd)
    except:
        print(datetime.now().strftime("%Y-%m-%d %H:%M:%S") + '\tCreate table failed')
        return False
    print(datetime.now().strftime("%Y-%m-%d %H:%M:%S") + "\tdb Connected!")
    conn.close()
       
def create_dir_not_exist(path):
    if not os.path.exists(path):
        print(datetime.now().strftime("%Y-%m-%d %H:%M:%S") + '\t' + path + ' not exitsts ,created!')
        os.mkdir(path)
    
def readinfo():
    conn = sqlite3.connect(abspath + '\\Res\\sgcc.db')
    try:
        # 1.创建游标对象
        cursor = conn.cursor()
        # 2.执行SQL操作
        # noinspection SqlResolve
        cursor.execute("select * from info order by id desc limit 0,1")
        data = cursor.fetchone()
        read_d={}
        col_name = [tuple[0] for tuple in cursor.description][1:6]
        if data is not None:
            read_db = list(data[1:6])
            return dict(zip(col_name,read_db))
        else:
            for k in col_name:
                read_d[k]=0
            return read_d
    except sqlite3.DatabaseError as error:
        # 4. 回滚数据库事物
        print(error)
        conn.rollback()
    finally:
        # 5. 关闭数据连接
        conn.close()
        
def insert_info(values):
    conn = sqlite3.connect(abspath + '\\Res\\sgcc.db')
    try:
        # 1.创建游标对象
        cursor = conn.cursor()
        columns ='\',\''.join(values.keys())
        placeholders = ':' + ', :'.join(values.keys())
        query = 'INSERT INTO info (\'%s\') VALUES (%s)' % (columns,placeholders)
        #insert into url(url)select 'bbs.aardio.com'where not exists(select urlid from url where url="bbs.aardio.com")
        cursor.execute(query, values)
        conn.commit()
    except sqlite3.DatabaseError as error:
        print(error)
        conn.rollback()
    finally:
        conn.close()    


if __name__ == '__main__':
    create_dir_not_exist(abspath + '\\Res')
    create_dir_not_exist(abspath + '\\Download')
    create_dir_not_exist(abspath + '\\PDF')
    create_table()
    dictnum=geturlnum()
    dicM=readinfo()
    addRows=0
    nNum={}
    for k in dictnum:
        addRows=addRows+int(dictnum[k])-int(dicM[k])
        nNum[k]=math.ceil((int(dictnum[k])-int(dicM[k]))/20)
    if addRows<=0:
        print(datetime.now().strftime("%Y-%m-%d %H:%M:%S") + "\t没有新的信息")
    else:
        Spider(nNum).run()
        dictnum['runtime']=datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        insert_info(dictnum)
        print(dictnum['runtime']+"\t写入excel")
        writexls(addRows)
        print(datetime.now().strftime("%Y-%m-%d %H:%M:%S") + "\t写入excel完成")
